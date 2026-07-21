from rest_framework import viewsets, status
from .models import Project, Task, Notification, FCMDevice
from .serializers import ProjectSerializer, TaskSerializer, CommentSerializer, NotificationSerializer
from rest_framework.decorators import action
from rest_framework.response import Response
from .services import shift_task_deadlines, generate_employee_excel
from apps.accounts.models import User
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from .permissions import IsManagerOrAdmin
from rest_framework.views import APIView
from django.db.models import Count, Q
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from .tasks import send_notification_email
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import permissions
from .models import Attachment, News, Vacation
from .serializers import AttachmentSerializer, NewsSerializer
from firebase_admin import messaging
import xml.etree.ElementTree as ET
import csv
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, date
from rest_framework.pagination import PageNumberPagination
from rest_framework.exceptions import PermissionDenied

User = get_user_model()

class TaskPagination(PageNumberPagination):
    page_size = 10  # Строго по 10 задач на страницу
    page_size_query_param = 'page_size' # Позволяет фронтенду менять размер при желании
    max_page_size = 100

    def paginate_queryset(self, queryset, request, view=None):
        # Если фронтенд явно просит отключить пагинацию, возвращаем None
        if request.query_params.get('no_page') == 'true':
            return None

        # В противном случае работает стандартная пагинация
        return super().paginate_queryset(queryset, request, view)

def notify_user(user, title, message, link=None, send_email=True):
    """
    Создает уведомление в БД, отправляет Email и Web Push.
    """
    if not user:
        return

    # 1. Создаем запись для интерфейса (колокольчик)
    Notification.objects.create(
        user=user,
        title=title,
        message=message
    )

    # 2. Отправляем письмо через Celery
    if send_email and getattr(user, 'email', None):
        try:
            send_notification_email.delay(user.email, title, message)
        except Exception:
            pass

    # 3. Отправляем Web Push через Firebase
    try:
        devices = FCMDevice.objects.filter(user=user)
        tokens = [device.registration_id for device in devices]

        if tokens:
            webpush_config = None
            if link:
                # Делаем ссылку абсолютной для ОС-пушей
                full_url = f"https://erp.stekufa.ru{link}" if link.startswith('/') else link
                webpush_config = messaging.WebpushConfig(
                    fcm_options=messaging.WebpushFCMOptions(link=full_url)
                )

            push_msg = messaging.MulticastMessage(
                notification=messaging.Notification(title=title, body=message),
                data={'link': link} if link else {},  # Отдаем фронтенду скрыто
                webpush=webpush_config,
                tokens=tokens,
            )
            messaging.send_each_for_multicast(push_msg)
    except Exception:
        pass


class DashboardOverduePagination(PageNumberPagination):
    page_size = 10  # Ровно 10 задач на страницу
    page_size_query_param = 'page_size'


def calculate_finish_date(start_date_str, duration_str):
    """
    Высчитывает дату дедлайна на основе даты начала и длительности MS Project.
    Пропускает выходные дни (субботу и воскресенье).
    """
    if not start_date_str or not duration_str or not duration_str.startswith('PT'):
        return start_date_str

    try:
        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')

        # Ищем часы и минуты с помощью регулярных выражений
        h_match = re.search(r'(\d+)H', duration_str)
        m_match = re.search(r'(\d+)M', duration_str)

        hours = int(h_match.group(1)) if h_match else 0
        minutes = int(m_match.group(1)) if m_match else 0

        # Переводим в рабочие дни (1 стандартный день = 8 часов)
        working_days = int((hours + (minutes / 60)) / 8)

        current_date = start_dt
        # Прибавляем дни, пропуская выходные (5 = суббота, 6 = воскресенье)
        while working_days > 0:
            current_date += timedelta(days=1)
            if current_date.weekday() < 5:
                working_days -= 1

        return current_date.strftime('%Y-%m-%d')
    except Exception:
        return start_date_str


class CanEditTaskPermission(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return True

        if view.action in ['add_comment', 'upload_files']:
            return True

        # === ПРОВЕРКА РОЛЕЙ ===
        is_boss = (
            getattr(request.user, 'role', '') in ['admin', 'director'] or
            request.user.is_superuser or
            (obj.project and obj.project.manager == request.user) or
            (obj.project and obj.project.visibility == 'selected' and obj.project.allowed_users.filter(id=request.user.id).exists()) or
            (obj.project and obj.project.visibility == 'all' and getattr(request.user, 'role', '') == 'manager')
        )
        is_assignee = (obj.assignee == request.user)
        # Добавляем проверку на исполнителя для прав редактирования:
        is_executor = (getattr(obj, 'executor', None) == request.user)
        is_participant = obj.participants.filter(id=request.user.id).exists()

        if view.action == 'hide':
            return is_boss or is_assignee or is_executor or is_participant

        if request.method in ['PATCH', 'PUT'] and (is_boss or is_assignee or is_executor or is_participant):
            return True

        return is_boss or is_assignee or is_executor


class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSerializer
    permission_classes = [IsAuthenticated, IsManagerOrAdmin]

    def perform_create(self, serializer):
        # Автоматически делаем текущего пользователя создателем/владельцем проекта
        serializer.save(owner=self.request.user)

    def get_queryset(self):
        user = self.request.user

        # Директора и Админы видят ВСЕ проекты (включая Приватные)
        if user.role in ['admin', ] or user.is_superuser:
            return Project.objects.all().distinct()

        # Для остальных собираем по правилам
        return Project.objects.filter(
            Q(owner=user) |  # Создатель
            Q(manager=user) |  # Ответственный за проект
            Q(visibility='all') |  # Видят все
            Q(visibility='department', owner__departments__in=user.departments.all()) |  # Отдел
            Q(visibility='selected', allowed_users=user)  # Несколько (добавленные пользователи)
        ).distinct()

    @action(detail=True, methods=['post'])
    def import_xml(self, request, pk=None):
        project_instance = self.get_object()
        xml_file = request.FILES.get('file')

        if not xml_file:
            return Response({"error": "Файл не предоставлен"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            ns = {'ns': root.tag.split('}')[0].strip('{')} if '}' in root.tag else {}
            prefix = 'ns:' if ns else ''

            wbs_to_task = {}  # Память для иерархии по коду WBS
            level_to_task = {}  # Резервная память для иерархии по уровню
            uid_to_task = {}  # Память для связей по UID
            tasks_with_dependencies = []
            created_count = 0

            # === ЭТАП 1: СОЗДАНИЕ ВСЕХ ЗАДАЧ ===
            for task_elem in root.findall(f'.//{prefix}Task', ns):
                task_id_elem = task_elem.find(f'{prefix}ID', ns)
                if task_id_elem is None or task_id_elem.text == '0':
                    continue

                name_elem = task_elem.find(f'{prefix}Name', ns)
                if name_elem is None or not name_elem.text:
                    continue

                name = name_elem.text

                uid_elem = task_elem.find(f'{prefix}UID', ns)
                task_uid = uid_elem.text if uid_elem is not None else None

                wbs_elem = task_elem.find(f'{prefix}WBS', ns)
                wbs = wbs_elem.text if wbs_elem is not None else None

                outline_level_elem = task_elem.find(f'{prefix}OutlineLevel', ns)
                outline_level = int(outline_level_elem.text) if outline_level_elem is not None else 1

                start_elem = task_elem.find(f'{prefix}Start', ns)
                finish_elem = task_elem.find(f'{prefix}Finish', ns)
                duration_elem = task_elem.find(f'{prefix}Duration', ns)
                notes_elem = task_elem.find(f'{prefix}Notes', ns)

                start_text = start_elem.text if start_elem is not None else None
                finish_text = finish_elem.text if finish_elem is not None else None
                duration_text = duration_elem.text if duration_elem is not None else None

                start_date = start_text.split('T')[0] if start_text else None
                finish_date = finish_text.split('T')[0] if finish_text else None

                # --- УМНЫЙ РАСЧЕТ ДАТ ---
                if not finish_date and start_date and duration_text:
                    finish_date = calculate_finish_date(start_date, duration_text)
                elif not finish_date and start_date:
                    finish_date = start_date
                elif not start_date and finish_date:
                    start_date = finish_date
                elif not start_date and not finish_date:
                    today_str = date.today().strftime('%Y-%m-%d')
                    start_date = today_str
                    finish_date = today_str

                description = notes_elem.text if notes_elem is not None else 'Импортировано из MS Project'

                # --- ДВОЙНАЯ ПОДСТРАХОВКА ИЕРАРХИИ ---
                parent_task = None
                if wbs and '.' in wbs:
                    parent_wbs = wbs.rsplit('.', 1)[0]
                    parent_task = wbs_to_task.get(parent_wbs)
                if not parent_task:
                    parent_task = level_to_task.get(outline_level - 1)

                new_task = Task.objects.create(
                    project=project_instance,
                    title=name,
                    description=description,
                    plan_start_date=start_date,
                    plan_end_date=finish_date,
                    status='new',
                    priority='medium',
                    assignee=request.user,
                    parent_task=parent_task
                )
                created_count += 1

                if wbs:
                    wbs_to_task[wbs] = new_task
                if task_uid:
                    uid_to_task[task_uid] = new_task

                level_to_task[outline_level] = new_task
                keys_to_delete = [k for k in level_to_task.keys() if k > outline_level]
                for k in keys_to_delete:
                    del level_to_task[k]

                # --- СЧИТЫВАЕМ СВЯЗИ ---
                predecessor_uids = []
                for link in task_elem.findall(f'{prefix}PredecessorLink', ns):
                    pred_uid_elem = link.find(f'{prefix}PredecessorUID', ns)
                    if pred_uid_elem is not None and pred_uid_elem.text:
                        predecessor_uids.append(pred_uid_elem.text)

                if predecessor_uids:
                    tasks_with_dependencies.append((new_task, predecessor_uids))

            # === ЭТАП 2: ПРОСТАВЛЯЕМ СВЯЗИ В БД ===
            for task_obj, pred_uids in tasks_with_dependencies:
                for p_uid in pred_uids:
                    parent_dependency = uid_to_task.get(p_uid)
                    if parent_dependency:
                        task_obj.dependencies.add(parent_dependency)

            return Response({"message": f"Успешно импортировано {created_count} задач со связями и вложенностью!"},
                            status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": f"Ошибка парсинга XML: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def export_excel(self, request, pk=None):
        """
        Выгрузка всех задач текущего проекта в красивый Excel-файл.
        """
        project = self.get_object()
        tasks = Task.objects.filter(project=project).order_by('plan_start_date')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Список задач"

        font_title = Font(name='Arial', size=14, bold=True, color='1E293B')
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_body = Font(name='Arial', size=10)

        fill_header = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        ws.merge_cells('A1:H1')
        ws['A1'] = f"Проект: {project.title}"
        ws['A1'].font = font_title
        ws.row_dimensions[1].height = 30
        ws.append([])

        headers = ["ID", "Название задачи", "Описание", "Исполнитель", "Статус", "Критичность", "Дата начала",
                   "Дедлайн"]
        ws.append(headers)

        ws.row_dimensions[3].height = 25
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        status_mapping = {'new': 'Новая', 'in_progress': 'В работе', 'completed': 'Завершена'}
        priority_mapping = {'low': 'Низкая', 'medium': 'Средняя', 'high': 'Высокая', 'critical': 'Критичная'}

        for task in tasks:
            assignee_name = "Не назначен"
            if task.assignee:
                assignee_name = task.assignee.get_full_name() or task.assignee.username

            row = [
                task.id,
                task.title,
                task.description or "",
                assignee_name,
                status_mapping.get(task.status, task.status),
                priority_mapping.get(task.priority, task.priority),
                task.plan_start_date.strftime('%Y-%m-%d') if task.plan_start_date else "",
                task.plan_end_date.strftime('%Y-%m-%d') if task.plan_end_date else ""
            ]
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=8):
            ws.row_dimensions[row[0].row].height = 20
            for cell in row:
                cell.font = font_body
                if cell.column in [1, 5, 6, 7, 8]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Project_{project.id}_Tasks.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    @action(detail=True, methods=['post'])
    def import_csv(self, request, pk=None):
        project_instance = self.get_object()
        csv_file = request.FILES.get('file')

        if not csv_file:
            return Response({"error": "Файл не предоставлен"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            file_bytes = csv_file.read()
            try:
                decoded_file = file_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                decoded_file = file_bytes.decode('windows-1251')

            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string, delimiter=',')

            for _ in range(4):
                next(reader, None)

            headers = next(reader, None)
            if not headers:
                return Response({"error": "Не удалось прочитать заголовки CSV"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                name_idx = headers.index('Наименование задач')
                start_idx = headers.index('Дата начала')
                end_idx = headers.index('Дата окончания')
                desc_idx = headers.index('Описание задачи')
                level_idx = headers.index('Уровень')
            except ValueError as e:
                return Response({"error": f"Не найдена колонка: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

            level_to_task = {}
            created_count = 0

            for row in reader:
                if not row or len(row) <= max(name_idx, level_idx):
                    continue

                name = row[name_idx].strip()
                if not name:
                    continue

                try:
                    outline_level = int(row[level_idx].strip())
                except ValueError:
                    outline_level = 1

                start_str = row[start_idx].strip()
                end_str = row[end_idx].strip()

                start_date = start_str if start_str else None
                finish_date = end_str if end_str else None

                if not finish_date and start_date:
                    finish_date = start_date
                elif not start_date and finish_date:
                    start_date = finish_date
                elif not start_date and not finish_date:
                    today_str = date.today().strftime('%Y-%m-%d')
                    start_date = today_str
                    finish_date = today_str

                description = row[desc_idx].strip() if len(row) > desc_idx else 'Импортировано из GanttPRO'
                parent_task = level_to_task.get(outline_level - 1)

                new_task = Task.objects.create(
                    project=project_instance,
                    title=name,
                    description=description,
                    plan_start_date=start_date,
                    plan_end_date=finish_date,
                    status='new',
                    priority='medium',
                    assignee=request.user,
                    parent_task=parent_task
                )
                created_count += 1

                level_to_task[outline_level] = new_task
                keys_to_delete = [k for k in level_to_task.keys() if k > outline_level]
                for k in keys_to_delete:
                    del level_to_task[k]

            return Response({"message": f"Успешно импортировано {created_count} задач!"},
                            status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": f"Ошибка обработки CSV: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def toggle_pin(self, request, pk=None):
        project = self.get_object()
        user = request.user

        if project.pinned_by.filter(id=user.id).exists():
            project.pinned_by.remove(user)
            is_pinned = False
        else:
            project.pinned_by.add(user)
            is_pinned = True

        return Response({'is_pinned': is_pinned})


class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated, CanEditTaskPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project']
    pagination_class = TaskPagination

    def get_queryset(self):
        user = self.request.user

        # 1. Оптимизация запросов к БД
        queryset = Task.objects.select_related(
            'project', 'assignee', 'parent_task'
        ).prefetch_related(
            'comments', 'attachments', 'dependencies', 'participants', 'hidden_for'
        )

        # 2. Фильтрация по проекту
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # 3. Фильтрация "Мои задачи"
        assigned_to_me = self.request.query_params.get('assigned_to_me')
        if assigned_to_me == 'true':
            queryset = queryset.filter(Q(assignee=user) | Q(executor=user) | Q(participants=user)).distinct()

        # 4. НОВОЕ: Фильтрация по статусу (для дашборда)
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # 5. Исключение скрытых задач
        if user.is_authenticated:
            queryset = queryset.exclude(hidden_for=user)

        return queryset.order_by('-created_at')

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        user = request.user
        today = date.today()

        # 1. Берем ту же оптимизацию, что и в get_queryset
        queryset = Task.objects.select_related(
            'project', 'assignee', 'parent_task'
        ).prefetch_related(
            'comments', 'attachments', 'dependencies', 'participants', 'hidden_for'
        )

        # 2. Оставляем только задачи, где юзер - исполнитель или участник
        queryset = queryset.filter(Q(assignee=user) | Q(participants=user))

        # 3. Фильтруем просроченные: дедлайн в прошлом, статус не завершен и не отложен
        queryset = queryset.filter(
            plan_end_date__lt=today
        ).exclude(
            status__in=['completed', 'delayed']
        )

        # 4. Исключаем скрытые задачи
        if user.is_authenticated:
            queryset = queryset.exclude(hidden_for=user)

        # 5. Сортируем так, чтобы самые "горящие" (старые дедлайны) были сверху
        queryset = queryset.distinct().order_by('plan_end_date')

        # 6. Применяем стандартную пагинацию DRF
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        """ Триггер 1: Создание новой задачи """
        task = serializer.save()

        if task.assignee and task.assignee != self.request.user:
            # Формируем ссылки для уведомления
            task_link = f"/task/{task.id}"
            full_url = f"https://erp.stekufa.ru{task_link}"

            notify_user(
                user=task.assignee,
                title="Новая задача",
                message=f"Вы назначены исполнителем новой задачи: {task.title}.\nПерейти к задаче: {full_url}",
                link=task_link
            )

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def hide(self, request, pk=None):
        """
        Скрытие задачи с доски Участника и автоматический лог в комментарии.
        """
        task = self.get_object()
        user = request.user

        task.hidden_for.add(user)

        target_status = request.data.get('status')
        if target_status:
            full_name = user.get_full_name() or user.username
            text = ""

            if target_status == 'in_progress':
                text = f"⚙️ {full_name} принял(а) задачу в работу"
            elif target_status == 'completed':
                text = f"✅ {full_name} завершил(а) свою часть работы"

            if text:
                serializer = CommentSerializer(data={'text': text})
                if serializer.is_valid():
                    serializer.save(task=task, author=user)

                    if task.assignee and task.assignee != user:
                        # Добавляем ссылки для ответственного
                        task_link = f"/task/{task.id}"
                        full_url = f"https://erp.stekufa.ru{task_link}"

                        notify_user(
                            user=task.assignee,
                            title="Обновление от участника",
                            message=f"{full_name} передвинул задачу '{task.title}'.\n{text}\nПерейти к задаче: {full_url}",
                            link=task_link
                        )

        return Response({'detail': 'Задача скрыта, лог сохранен'}, status=status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        task = self.get_object()
        user = request.user
        project = task.project

        is_boss = (
                user.role in ['admin', 'director'] or
                user.is_superuser or
                user == project.owner or
                user == project.manager or
                (project.visibility == 'selected' and project.allowed_users.filter(id=user.id).exists()) or
                (project.visibility == 'all' and getattr(user, 'role', '') == 'manager')
        )
        is_assignee = (user == task.assignee)
        is_participant = task.participants.filter(id=user.id).exists()

        data_to_save = request.data.copy()

        # === ЧИТАЕМ ФЛАГ С ФРОНТЕНДА ===
        is_personal_update = str(request.data.get('personal_only', '')).lower() == 'true'

        # Если Участник перетащил карточку на Канбане (personal_only=True)
        # ИЛИ это просто рядовой Участник (не босс/исполнитель) меняет статус через модалку
        if is_participant and (is_personal_update or (not is_assignee and not is_boss)):
            status_data = data_to_save.get('status')
            if status_data:
                from .models import TaskUserStatus, Notification

                TaskUserStatus.objects.update_or_create(
                    user=user,
                    task=task,
                    defaults={'status': status_data}
                )

                full_name = user.get_full_name() or user.username
                status_label = dict(Task.STATUS_CHOICES).get(status_data, status_data) if hasattr(Task,
                                                                                                  'STATUS_CHOICES') else status_data
                text = f"🏃‍♂️ {full_name} перевел(а) свою часть работы в статус '{status_label}'"

                from .serializers import CommentSerializer
                serializer_comment = CommentSerializer(data={'text': text})
                if serializer_comment.is_valid():
                    serializer_comment.save(task=task, author=user)

                if task.assignee and task.assignee != user:
                    task_link = f"/task/{task.id}"
                    full_url = f"https://erp.stekufa.ru{task_link}"
                    notify_user(
                        user=task.assignee,
                        title="Участник обновил статус",
                        message=f"{text} в задаче '{task.title}'.\nПерейти к задаче: {full_url}",
                        link=task_link
                    )

                serializer = self.get_serializer(task)
                return Response(serializer.data)

        # --- СТАНДАРТНОЕ ГЛОБАЛЬНОЕ ОБНОВЛЕНИЕ ---
        # Сюда провалятся Боссы и Исполнители, если они не отправляли флаг personal_only
        if not is_boss:
            if is_assignee:
                allowed_keys = ['status', 'delay_reason', 'actual_end_date']
                data_to_save = {k: v for k, v in data_to_save.items() if k in allowed_keys}
            else:
                return Response(
                    {'detail': 'У вас нет прав на редактирование параметров этой задачи.'},
                    status=status.HTTP_403_FORBIDDEN
                )

        serializer = self.get_serializer(task, data=data_to_save, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(task, '_prefetched_objects_cache', None):
            task._prefetched_objects_cache = {}

        return Response(serializer.data)

    def perform_update(self, serializer):
        """ Триггер 2: Обновление задачи (смена статуса/сроков) """
        old_task = self.get_object()
        old_status = old_task.status
        old_assignee = old_task.assignee

        task = serializer.save()

        # Заранее формируем ссылки
        task_link = f"/task/{task.id}"
        full_url = f"https://erp.stekufa.ru{task_link}"

        if old_status != task.status:
            if task.status == 'completed' and task.project.manager:
                notify_user(
                    user=task.project.manager,
                    title="Задача завершена",
                    message=f"Задача '{task.title}' была завершена исполнителем.\nПерейти к задаче: {full_url}",
                    link=task_link
                )

            if task.assignee and self.request.user != task.assignee:
                notify_user(
                    user=task.assignee,
                    title="Статус задачи изменен",
                    message=f"Статус вашей задачи '{task.title}' изменен на '{task.get_status_display()}'.\nПерейти к задаче: {full_url}",
                    link=task_link
                )

        # === ПРОВЕРКА СМЕНЫ ИСПОЛНИТЕЛЯ ===
        if old_assignee != task.assignee:
            # 1. Формируем имена для комментария
            old_name = old_assignee.get_full_name() or old_assignee.username if old_assignee else "Не назначен"
            new_name = task.assignee.get_full_name() or task.assignee.username if task.assignee else "Не назначен"

            # 2. Создаем авто-комментарий
            from .models import Comment
            Comment.objects.create(
                task=task,
                author=self.request.user,
                text=f"🔄 В задаче сменился ответственный. Был: {old_name} стал: {new_name}"
            )

            # 3. Отправляем уведомление новому исполнителю (если он есть и это не сам автор изменений)
            if task.assignee and task.assignee != self.request.user:
                notify_user(
                    user=task.assignee,
                    title="Новое назначение",
                    message=f"Вы были назначены ответственным за задачу: {task.title}.\nПерейти к задаче: {full_url}",
                    link=task_link
                )

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def dashboard_metrics(self, request):
        user = request.user
        today = date.today()

        # Теперь мы ищем задачи, где пользователь - Исполнитель ИЛИ Участник
        base_query = Q(assignee=user) | Q(participants=user) | Q(executor=user)

        metrics = Task.objects.filter(base_query).distinct().aggregate(
            total=Count('id'),
            new_tasks=Count('id', filter=Q(status='new')),
            in_progress=Count('id', filter=Q(status='in_progress')),
            completed=Count('id', filter=Q(status='completed')),
            # Просроченные: дедлайн в прошлом, и статус не Завершен/Отсрочка
            overdue_count=Count('id', filter=Q(plan_end_date__lt=today) & ~Q(status__in=['completed', 'delayed']))
        )
        return Response(metrics)

    @action(detail=False, methods=['get'])
    def overdue(self, request):
        user = request.user
        today = date.today()

        # Ищем просроченные задачи для Исполнителя и Участника
        base_query = Q(assignee=user) | Q(participants=user)

        qs = Task.objects.filter(
            base_query,
            plan_end_date__lt=today
        ).exclude(
            status__in=['completed', 'delayed']
        ).select_related(
            'project', 'assignee'
        ).distinct().order_by('plan_end_date')

        paginator = DashboardOverduePagination()
        page = paginator.paginate_queryset(qs, request)
        serializer = self.get_serializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    @action(detail=True, methods=['post'])
    def shift_deadlines(self, request, pk=None):
        task = self.get_object()
        days_delta = request.data.get('days', 0)
        cascade = request.data.get('cascade', True)

        try:
            days_delta = int(days_delta)
        except ValueError:
            return Response({"error": "Поле days должно быть целым числом"}, status=status.HTTP_400_BAD_REQUEST)

        shift_task_deadlines(task=task, days_delta=days_delta, cascade=cascade)

        task.refresh_from_db()
        serializer = self.get_serializer(task)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def add_comment(self, request, pk=None):
        """ Триггер 3: Новый комментарий """
        task = self.get_object()
        serializer = CommentSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save(task=task, author=request.user)

            author_name = request.user.get_full_name() or request.user.username
            task_link = f"/task/{task.id}"
            full_url = f"https://erp.stekufa.ru{task_link}"

            # 1. Собираем уникальный список всех, кто причастен к задаче
            users_to_notify = set()

            if task.assignee:
                users_to_notify.add(task.assignee)

            if getattr(task, 'executor', None):
                users_to_notify.add(task.executor)

            for participant in task.participants.all():
                users_to_notify.add(participant)

            # 2. Исключаем из рассылки автора комментария
            if request.user in users_to_notify:
                users_to_notify.remove(request.user)

            # 3. Отправляем уведомления всем собранным пользователям
            for target_user in users_to_notify:
                notify_user(
                    user=target_user,
                    title="💬 Новый комментарий",
                    message=f"{author_name} оставил(а) комментарий в задаче '{task.title}'.\nПерейти к задаче: {full_url}",
                    link=task_link,
                    send_email=False,
                )

            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_files(self, request, pk=None):
        task = self.get_object()
        if 'file' not in request.FILES:
            return Response({"error": "Файл не найден"}, status=status.HTTP_400_BAD_REQUEST)
        file_obj = request.FILES['file']

        attachment = Attachment.objects.create(
            task=task,
            file=file_obj,
            uploaded_by=request.user,
        )
        serializer = AttachmentSerializer(attachment)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class AttachmentViewSet(viewsets.ModelViewSet):
    queryset = Attachment.objects.all()
    serializer_class = AttachmentSerializer


class EmployeeReportView(APIView):
    def get(self, request, user_id):
        try:
            employee = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"error": "Сотрудник не найден"}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        tasks = Task.objects.filter(assignee=employee).order_by('-id')

        stats = tasks.aggregate(
            total_tasks=Count('id'),
            completed_tasks=Count('id', filter=Q(status='completed')),
            overdue_tasks=Count('id', filter=~Q(status='completed') & Q(plan_end_date__lt=today)),
            in_progress_tasks=Count('id', filter=Q(status='in_progress'))
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет по сотруднику"

        font_title = Font(name='Arial', size=14, bold=True)
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A1:E1')
        ws[
            'A1'] = f"Отчет по сотруднику: {employee.get_full_name() or employee.username} ({employee.position or 'Должность не указана'})"
        ws['A1'].font = font_title
        ws.row_dimensions[1].height = 25

        ws.append(["Всего задач:", stats['total_tasks']])
        ws.append(["В работе:", stats['in_progress_tasks']])
        ws.append(["Завершено:", stats['completed_tasks']])
        ws.append(["Просрочено:", stats['overdue_tasks']])
        ws.append([])

        headers = ["ID", "Название задачи", "Статус", "Критичность", "Дедлайн"]
        ws.append(headers)

        header_row_idx = ws.max_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        status_map = {'new': 'Новая', 'in_progress': 'В работе', 'completed': 'Завершена'}
        priority_map = {'low': 'Низкая', 'medium': 'Средняя', 'high': 'Высокая', 'critical': 'Критичная'}

        for task in tasks:
            ws.append([
                task.id,
                task.title,
                status_map.get(task.status, task.status),
                priority_map.get(task.priority, task.priority),
                task.plan_end_date.strftime('%Y-%m-%d') if task.plan_end_date else "Не указан"
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 12)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Employee_{employee.id}_Report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response


class ProjectReportView(APIView):
    def get(self, request, project_id):
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({"error": "Проект не найден"}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        tasks = project.tasks.all()

        project_stats = tasks.aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            in_progress=Count('id', filter=Q(status='in_progress')),
            overdue=Count('id', filter=~Q(status='completed') & Q(plan_end_date__lt=today))
        )

        employee_stats = tasks.values(
            'assignee__id', 'assignee__first_name', 'assignee__last_name'
        ).annotate(
            total_tasks=Count('id'),
            overdue_tasks=Count('id', filter=~Q(status='completed') & Q(plan_end_date__lt=today))
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет по проекту"

        font_title = Font(name='Arial', size=14, bold=True)
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        fill_header = PatternFill(start_color='0F766E', end_color='0F766E', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A1:D1')
        ws['A1'] = f"Проект: {project.title} (Статус: {project.get_status_display()})"
        ws['A1'].font = font_title
        ws.row_dimensions[1].height = 25

        ws.append(["Всего задач:", project_stats['total']])
        ws.append(["В работе:", project_stats['in_progress']])
        ws.append(["Завершено:", project_stats['completed']])
        ws.append(["Просрочено:", project_stats['overdue']])
        ws.append([])

        ws.append(["Статистика по исполнителям"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

        emp_headers = ["ID", "Сотрудник", "Всего задач", "Просрочено"]
        ws.append(emp_headers)

        for col_num, header in enumerate(emp_headers, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for emp in employee_stats:
            name = f"{emp['assignee__first_name'] or ''} {emp['assignee__last_name'] or ''}".strip()
            ws.append([
                emp['assignee__id'] or "-",
                name if name else ("Не назначен" if not emp['assignee__id'] else f"Пользователь {emp['assignee__id']}"),
                emp['total_tasks'],
                emp['overdue_tasks']
            ])

        ws.append([])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 14)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Project_{project.id}_Report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response


class NewsPagination(PageNumberPagination):
    page_size = 3


class NewsViewSet(viewsets.ModelViewSet):
    queryset = News.objects.all()
    serializer_class = NewsSerializer
    pagination_class = NewsPagination
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        user = self.request.user
        if user.role in ['admin', 'director'] or user.is_superuser or getattr(user, 'can_post_news', False):
            serializer.save(author=user)
        else:
            raise PermissionDenied("У вас нет прав публиковать новости.")

    def destroy(self, request, *args, **kwargs):
        user = request.user
        if user.role in ['admin', 'director'] or user.is_superuser or getattr(user, 'can_post_news', False):
            return super().destroy(request, *args, **kwargs)
        raise PermissionDenied("У вас нет прав удалять новости.")


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

    @action(detail=True, methods=['post'])
    def mark_as_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'status': 'Уведомление прочитано'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def mark_all_as_read(self, request):
        updated_count = self.get_queryset().filter(is_read=False).update(is_read=True)
        return Response({'status': f'{updated_count} уведомлений прочитано'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def save_fcm_token(self, request):
        token = request.data.get('token')
        if token:
            FCMDevice.objects.get_or_create(user=request.user, registration_id=token)
            return Response({'status': 'Токен успешно сохранен'}, status=status.HTTP_200_OK)
        return Response({'error': 'Токен не предоставлен'}, status=status.HTTP_400_BAD_REQUEST)


class VacationViewSet(viewsets.ModelViewSet):
    queryset = Vacation.objects.all()
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        csv_file = request.FILES.get('file')
        if not csv_file:
            return Response({"error": "Файл не предоставлен"}, status=400)

        file_bytes = csv_file.read()
        try:
            decoded = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = file_bytes.decode('windows-1251')

        reader = csv.reader(io.StringIO(decoded), delimiter=',')
        next(reader, None)  # Пропускаем строку заголовков

        created_count = 0
        not_found_users = []
        double_users = []
        bad_dates = []

        # Внутренняя умная функция перевода "02.02.26" -> "2026-02-02"
        def parse_date(d_str):
            d_str = d_str.strip().replace('/', '.')
            for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(d_str, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    continue
            return None

        for row in reader:
            if not row or len(row) < 4:
                continue

            fio = row[1].strip()
            raw_start = row[2].strip()
            raw_end = row[3].strip()

            if not fio or not raw_start or not raw_end:
                continue

            # Преобразуем даты
            start_str = parse_date(raw_start)
            end_str = parse_date(raw_end)

            if not start_str or not end_str:
                bad_dates.append(f"{fio} ({raw_start} - {raw_end})")
                continue

            # Разбираем ФИО
            parts = fio.split()
            if len(parts) < 2:
                not_found_users.append(fio)
                continue

            last_name = parts[0]
            first_name = parts[1]

            matched = User.objects.filter(
                last_name__iexact=last_name,
                first_name__iexact=first_name
            )

            if matched.count() == 0:
                not_found_users.append(fio)
                continue
            elif matched.count() > 1:
                double_users.append(fio)
                continue

            user_obj = matched.first()

            Vacation.objects.create(
                user=user_obj,
                start_date=start_str,
                end_date=end_str
            )
            created_count += 1

        response_data = {
            "message": f"Успешно импортировано {created_count} отпусков!",
            "not_found": not_found_users,
            "doubles": double_users
        }

        # Если кадровик где-то ввел совсем космическую дату вроде "32.13.26"
        if bad_dates:
            response_data["bad_dates"] = bad_dates

        return Response(response_data, status=200)