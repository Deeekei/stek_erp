from datetime import timedelta
from django.db import transaction
from .models import Task
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO


@transaction.atomic
def shift_task_deadlines(task: Task, days_delta: int, cascade: bool = True):
    if days_delta == 0:
        return
    if task.plan_start_date:
        task.plan_start_date += timedelta(days=days_delta)

    if task.plan_end_date:
        task.plan_end_date += timedelta(days=days_delta)

    task.save(update_fields=['plan_start_date', 'plan_end_date'])

    if cascade:
        dependent_tasks = task.dependencies.all()

        for dep_task in dependent_tasks:
            shift_task_deadlines(dep_task, days_delta, cascade=True)


def generate_employee_excel(employee, stats):
    """
    Создает Excel-документ со сводкой по сотруднику и списком его задач.
    Возвращает файл в виде байтового потока (BytesIO).
    """
    # 1. Создаем новую книгу Excel и активный лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по сотруднику"

    # 2. Записываем общую статистику (шапка отчета)
    ws.append(["Сводка по сотруднику:", employee.get_full_name()])
    ws.append(["Всего задач:", stats['total_tasks']])
    ws.append(["Завершено:", stats['completed_tasks']])
    ws.append(["В работе:", stats['in_progress_tasks']])
    ws.append(["Просрочено:", stats['overdue_tasks']])
    ws.append([])  # Пустая строка для отступа

    # 3. Создаем заголовки для таблицы задач
    headers = ["Проект", "Задача", "Статус", "План. завершение", "Факт. завершение", "Причина просрочки"]
    ws.append(headers)

    # Делаем заголовки таблицы жирными
    header_row = ws.max_row
    for col in range(1, 7):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 4. Получаем все задачи сотрудника и заполняем таблицу
    # select_related('project') делает SQL-запрос быстрее, сразу подтягивая данные проекта
    tasks = Task.objects.filter(assignee=employee).select_related('project').order_by('-created_at')

    for task in tasks:
        ws.append([
            task.project.title,
            task.title,
            task.get_status_display(),  # Получает красивое русское название статуса
            task.plan_end_date.strftime("%d.%m.%Y") if task.plan_end_date else "-",
            task.actual_end_date.strftime("%d.%m.%Y") if task.actual_end_date else "-",
            task.delay_reason or "-"
        ])

    # 5. Сохраняем результат в виртуальный файл (в оперативной памяти)
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Возвращаем "курсор" в начало файла, чтобы его можно было прочитать

    return output


